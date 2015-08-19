# coding=utf-8
import collections
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponse

from ServicioSocial.models import *
from ServicioSocial.forms import *







# Register your models here.


class UserProfileInline(admin.StackedInline):
    model = UserProfile


class UsuariosInLine(admin.TabularInline):
    model = DetalleEspera
    form = AddDetalleEsperaForm
    extra = 0

    readonly_fields = ['email', 'first_name', 'last_name', 'carrera', 'telefono', 'facebook', 'matricula', 'semestre',
                       'fecha_registro']

    def email(self, instance):
        return instance.usuario.email

    def first_name(self, instance):
        return instance.usuario.first_name

    def last_name(self, instance):
        return instance.usuario.last_name

    def carrera(self, instance):
        return instance.usuario.userprofile.carrera.iniciales

    def telefono(self, instance):
        return instance.usuario.userprofile.telefono

    def facebook(self, instance):
        return instance.usuario.userprofile.facebook

    def matricula(self, instance):
        return instance.usuario.userprofile.matricula

    def semestre(self, instance):
        return instance.usuario.userprofile.semestre


class CustomUserAdmin(UserAdmin):
    inlines = (UserProfileInline, )
    list_display = ('username', 'first_name', 'last_name', 'email', )

    add_fieldsets = (
        (None, {'fields': ('username', 'password1', 'password2')}),
        (('Personal info'), {'fields': ('first_name', 'last_name', 'email')}),
        (('Permissions'), {'fields': ('is_active', 'is_staff', 'is_superuser',
                                      'groups', 'user_permissions')}),
        (('Important dates'), {'fields': ('last_login', 'date_joined')}),
    )


class UsuarioGrupoInLine(admin.TabularInline):
    model = DetalleInscripcion
    extra = 0
    readonly_fields = ['email', 'first_name', 'last_name', 'carrera', 'telefono', 'facebook', 'matricula', 'semestre']

    def email(self, instance):
        return instance.usuario.email

    def first_name(self, instance):
        return instance.usuario.first_name

    def last_name(self, instance):
        return instance.usuario.last_name

    def carrera(self, instance):
        return instance.usuario.userprofile.carrera.iniciales

    def telefono(self, instance):
        return instance.usuario.userprofile.telefono

    def facebook(self, instance):
        return instance.usuario.userprofile.facebook

    def matricula(self, instance):
        return instance.usuario.userprofile.matricula

    def semestre(self, instance):
        return instance.usuario.userprofile.semestre


    email.short_description = 'Email'
    carrera.short_description = 'Carrera'


class HorarioInLine(admin.StackedInline):
    model = Horario
    extra = 0


class GrupoInLine(admin.StackedInline):
    model = Grupo
    extra = 1


def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ListasEspera.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Lista de Espera"

    row_num = 0

    columns = [
        (u"Grupo", 30),
        (u"Nombre", 30),
        (u"Apellidos", 30),
        (u"Aprobado", 30),
        (u"Email", 30),
        (u"Carrera", 30),
        (u"Telefono", 30),
        (u"Facebook", 30),
        (u"Matricula", 30),
        (u"Semestre", 30),
        (u"Fecha de Registro", 30),

    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    # Recibe un modelo de ListaEspera

    for obj in queryset:

        try:
            print obj
            for detalleEspera in obj.detalleespera_set.all():

                row_num += 1
                row = [
                    detalleEspera.__str__(),
                    detalleEspera.usuario.first_name,
                    detalleEspera.usuario.last_name,
                    detalleEspera.aprobado,
                    detalleEspera.usuario.email,
                    detalleEspera.usuario.userprofile.carrera.iniciales,
                    detalleEspera.usuario.userprofile.telefono,
                    detalleEspera.usuario.userprofile.facebook,
                    detalleEspera.usuario.userprofile.matricula,
                    detalleEspera.usuario.userprofile.semestre,
                    detalleEspera.fecha_registro
                ]

                for col_num in xrange(len(row)):
                    c = ws.cell(row=row_num + 1, column=col_num + 1)
                    c.value = row[col_num]
                    c.style.alignment.wrap_text = True

        except Exception as e:
            print e

    wb.save(response)
    return response


export_xlsx.short_description = u"Exportar a Excel"


def export_xlsx_duplicates(modeladmin, request, queryset):

    usuarios_proyecto = {}

    for listaEspera in queryset:
        usuarios_proyecto[listaEspera.grupo] = listaEspera.usuarios.all()

    all_lists = list(usuarios_proyecto.values())
    print all_lists
    common_links = set(all_lists[0]).intersection(*all_lists[1:])

    print common_links


export_xlsx_duplicates.short_description = u"Encontrar duplicados y exportar a Excel"





class ListaEsperaAdmin(admin.ModelAdmin):
    inlines = (UsuariosInLine, )
    actions = [export_xlsx, ]


class GrupoAdmin(admin.ModelAdmin):
    inlines = [HorarioInLine, UsuarioGrupoInLine]


class ProyectoAdmin(admin.ModelAdmin):
    inlines = (GrupoInLine, )
    fieldsets = (
        ('Descripcion de proyecto', {
            'fields': ('nombre', 'administrador', 'descripcion', 'numero_horas', 'ubicacion', 'fecha_inicio_proyecto',
                       'fecha_termino_proyecto')
        }),
        ('Fechas de registro', {
            'fields': ('fecha_registro_inicio', 'fecha_registro_fin',)
        }),
        ('Perfil', {
            'fields': ('requisitos', 'semestre_maximo', 'semestre_minimo', 'carreras')
        }),
    )

    list_display = (
        'nombre',
        'numero_horas',
        'fecha_inicio_proyecto',
        'fecha_termino_proyecto',
        'ubicacion',
        'administrador',
    )


admin.site.register(Ubicacion)
admin.site.register(Carrera)
admin.site.register(Grupo, GrupoAdmin)
admin.site.register(Proyecto, ProyectoAdmin)
admin.site.register(ListaEspera, ListaEsperaAdmin)
admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)
admin.site.register(Dia)